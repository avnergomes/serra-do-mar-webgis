#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Atlas do montanhismo paranaense: parse "Lista de Rotas de Escalada.xlsx" into
routes.json (full fidelity, every sheet) and routes.geojson (map scope: the Serra
do Mar / PR mountain crags, geolocated against peaks.geojson).

The workbook is hand-kept, so each sheet has its own shape. Three quirks drive
most of the code here:

  1. Grade cells are sparse: a grade is written once and applies to every row
     below it until the next one. Read naively you get ~2 grades per sheet.
  2. Puruna and Macarrao lay their sectors out side by side in column blocks
     rather than stacking them.
  3. Sheets carry running-total rows and repeated header rows mid-table.

Grades come in two families (Brazilian free and French sport) that do not share
a numeric axis, so we classify into four bands instead of faking a common float.
Stdlib + openpyxl.
"""
import json, os, re, sys, unicodedata
from collections import Counter, OrderedDict

import openpyxl

XLSX = "Lista de Rotas de Escalada.xlsx"
PEAKS = "peaks.geojson"
OUT_JSON = "routes.json"
OUT_GEO = "routes.geojson"

# --------------------------------------------------------------------------
# grade classification
# --------------------------------------------------------------------------

# Brazilian free-climbing scale, easiest first.
BR_SCALE = ["I", "II", "IIsup", "III", "IIIsup", "IV", "IVsup", "V", "Vsup",
            "VI", "VIsup", "VIIa", "VIIb", "VIIc", "VIIIa", "VIIIb", "VIIIc",
            "IXa", "IXb", "IXc", "Xa", "Xb", "Xc"]
# French sport scale, easiest first. The sheets mix the letter forms with the bare
# and "sup" forms Brazilians write ("6", "6sup", "9"), so those sit in the ordering
# too: bare N reads as the bottom of its range, Nsup as the top.
FR_SCALE = ["3", "4", "5", "5sup", "6", "6a", "6b", "6c", "6sup",
            "7", "7a", "7b", "7c", "7sup", "8", "8a", "8b", "8c",
            "9", "9a", "9b", "9c", "10a", "10b"]

# Band cutoffs, by index into the scale above. The two scales are aligned at the
# band level only (BR VIIa ~ FR 7a, BR V ~ FR 6a): close enough to colour a map,
# not precise enough to publish as a conversion table.
BANDS = ["iniciante", "intermediario", "avancado", "elite"]
BR_BANDS = {"iniciante": ("I", "IVsup"), "intermediario": ("V", "VIsup"),
            "avancado": ("VIIa", "VIIc"), "elite": ("VIIIa", "Xc")}
FR_BANDS = {"iniciante": ("3", "5sup"), "intermediario": ("6", "6sup"),
            "avancado": ("7", "7sup"), "elite": ("8", "10b")}
# Hueco V-scale, used by the two boulder sheets. Banded on its own axis: a V5
# boulder problem and a VIIa route are not the same thing, but both read as the
# start of "hard for this discipline".
V_BANDS = [(1, "iniciante"), (4, "intermediario"), (7, "avancado"), (99, "elite")]

BAND_LABEL = {"iniciante": "Iniciante", "intermediario": "Intermediário",
              "avancado": "Avançado", "elite": "Elite"}
BAND_COLOR = {"iniciante": "#52b788", "intermediario": "#ffb703",
              "avancado": "#e76f51", "elite": "#9d4edd"}


def _band_of(token, scale, bands):
    if token not in scale:
        return None
    i = scale.index(token)
    for band, (lo, hi) in bands.items():
        if scale.index(lo) <= i <= scale.index(hi):
            return band
    return None


def _norm_grade_text(s):
    """'IV sup' -> 'IVsup', '6° sup' -> '6sup'. Also strips the degree marker."""
    s = str(s).strip()
    s = s.replace("º", "°")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"(?i)\b(\w+)\s+sup\b", r"\1sup", s)
    return s


# Free-climbing token inside a compound grade like "VIsup A2 E4 M1" or "6 VIIc (A1) E3".
BR_TOKEN = re.compile(r"\b(X[abc]|IX[abc]|VIII[abc]|VII[abc]|VIsup|VI|Vsup|V|IVsup|IV|IIIsup|III|IIsup|II|I)(?![\wa-z])")
FR_TOKEN = re.compile(r"(?<![\w.])(10[ab]|9[abc]?|8[abc]?|7(?:sup|[abc])?|6(?:sup|[abc])?|5sup|5|4|3)(?![\w])")
AID_TOKEN = re.compile(r"\bA([0-5])\+?\b")
EXP_TOKEN = re.compile(r"\bE([1-6])\b")
MIXED_TOKEN = re.compile(r"\bM([1-5])\b")
APPROACH_TOKEN = re.compile(r"\bD([1-5])\b")
# Leading "6°" / "5° sup" is the overall route grade (grau global), not the crux move.
GLOBAL_TOKEN = re.compile(r"^(\d)°\s*(sup)?")
# Boulder V-scale, including the hedged forms the sheets use: "V6?", "V7/8".
V_TOKEN = re.compile(r"^V(\d+)")


def parse_vscale(raw):
    """'V7/8' -> {raw:'V7/8', v:7, band:'avancado'}. '?' marks an unconfirmed grade."""
    if raw is None:
        return None
    txt = str(raw).strip()
    m = V_TOKEN.match(txt)
    if not m:
        return None
    v = int(m.group(1))
    out = {"raw": txt, "v": v}
    for hi, band in V_BANDS:
        if v <= hi:
            out["band"] = band
            break
    if "?" in txt:
        out["incerto"] = True
    return out


def parse_grade(raw, prefer="br"):
    """Pull structure out of a free-form grade cell.

    Returns a dict with the raw text, the crux free token, its scale and band,
    plus aid / exposure / mixed / approach components when present. Returns None
    for empty or unknown cells ('?', '???').
    """
    if raw is None:
        return None
    txt = _norm_grade_text(raw)
    if not txt or txt.strip("?") == "":
        return None

    out = {"raw": txt}

    g = GLOBAL_TOKEN.search(txt)
    if g:
        out["global"] = g.group(1) + ("sup" if g.group(2) else "")
        txt_rest = txt[g.end():]
    else:
        txt_rest = txt

    # A parenthesised grade is the aid-free (livre) variant; keep it aside so it
    # does not shadow the route's own crux grade.
    variant = None
    m = re.search(r"\(([^)]*)\)", txt_rest)
    if m:
        variant = m.group(1).strip()
        txt_main = txt_rest[:m.start()] + " " + txt_rest[m.end():]
    else:
        txt_main = txt_rest

    free, scale = None, None
    if prefer == "br":
        mb = BR_TOKEN.search(txt_main)
        if mb:
            free, scale = mb.group(1), "br"
        else:
            mf = FR_TOKEN.search(txt_main)
            if mf:
                free, scale = mf.group(1), "fr"
    else:
        mf = FR_TOKEN.search(txt_main)
        if mf:
            free, scale = mf.group(1), "fr"
        else:
            mb = BR_TOKEN.search(txt_main)
            if mb:
                free, scale = mb.group(1), "br"

    if free:
        out["free"] = free
        out["scale"] = scale
        band = _band_of(free, BR_SCALE if scale == "br" else FR_SCALE,
                        BR_BANDS if scale == "br" else FR_BANDS)
        if band:
            out["band"] = band
    elif "global" in out:
        # Only a global grade (e.g. "4° sup"): band it on the BR axis by roman degree.
        approx = {"1": "III", "2": "III", "3": "IV", "4": "V", "5": "Vsup",
                  "6": "VIsup", "7": "VIIb"}.get(out["global"][0])
        if approx:
            out["band"] = _band_of(approx, BR_SCALE, BR_BANDS)
            out["band_from_global"] = True

    if variant:
        out["variant"] = variant
    a = AID_TOKEN.search(txt_main) or (AID_TOKEN.search(variant) if variant else None)
    if a:
        out["aid"] = a.group(0)
    e = EXP_TOKEN.search(txt_main)
    if e:
        out["exposure"] = e.group(0)
    mm = MIXED_TOKEN.search(txt_main)
    if mm:
        out["mixed"] = mm.group(0)
    dd = APPROACH_TOKEN.search(txt_main)
    if dd:
        out["approach"] = dd.group(0)
    return out


# --------------------------------------------------------------------------
# sheet reading helpers
# --------------------------------------------------------------------------

def rows_of(ws, limit=None):
    """Non-empty rows as lists, capped so a sheet with a bogus max_row
    (Puruna reports 1048576) does not stall the read."""
    out = []
    for r in ws.iter_rows(max_row=limit or min(ws.max_row, 3000), values_only=True):
        if any(c is not None and str(c).strip() != "" for c in r):
            out.append(list(r))
    return out


def cell(row, i):
    if i is None or i >= len(row):
        return None
    v = row[i]
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def is_header_row(row, words):
    joined = " ".join(str(c).lower() for c in row if c is not None)
    return any(w in joined for w in words)


def as_height(v):
    """'150 m' -> 150 ; 16 -> 16 ; junk -> None."""
    if v is None:
        return None
    s = str(v).strip().lower().replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m:
        return None
    n = float(m.group(1))
    return int(n) if n == int(n) else n


def as_year(v):
    if v is None:
        return None
    m = re.search(r"(1[89]\d\d|20\d\d)", str(v))
    return int(m.group(1)) if m else None


def slug(s):
    """Accent- and case-insensitive key for matching via names across sheets."""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


HEADER_WORDS = ["grau", "gradua", "via", "setor", "montanha", "boulder", "local"]


# --------------------------------------------------------------------------
# per-sheet parsers
# --------------------------------------------------------------------------

def parse_simple(ws, crag, cols, tipo="via", prefer="br", ffill=(0,)):
    """Sheets that stack rows with sparse (merged) grade cells."""
    out = []
    carry = {}
    for row in rows_of(ws):
        if is_header_row(row[:3], ["grau", "gradua"]):
            continue
        for i in ffill:
            v = cell(row, i)
            if v:
                carry[i] = v
        via = cell(row, cols["via"])
        if not via or via == "?":
            continue
        rec = {"via": via, "crag": crag, "tipo": tipo}
        gi = cols.get("grau")
        graw = cell(row, gi) or carry.get(gi)
        g = parse_grade(graw, prefer=prefer)
        if g:
            rec["grau"] = g
        for key in ("setor", "montanha", "bloco", "tecnica", "material"):
            if key in cols:
                v = cell(row, cols[key]) or (carry.get(cols[key]) if cols[key] in ffill else None)
                if v:
                    rec[key] = v
        if "altura" in cols:
            h = as_height(cell(row, cols["altura"]))
            if h:
                rec["altura"] = h
        if "vscale" in cols:
            v = parse_vscale(cell(row, cols["vscale"]) or carry.get(cols["vscale"]))
            if v:
                rec["vscale"] = v
        out.append(rec)
    return out


def parse_blocks(ws, crag, blocks, prefer="fr", tipo="via"):
    """Puruna / Macarrao: sectors laid out side by side in column blocks."""
    out = []
    for gi, vi, setor in blocks:
        carry = None
        for row in rows_of(ws):
            if is_header_row(row[:2], ["grau"]) or is_header_row(row[:1], ["setor", "frente", "fundos"]):
                continue
            g = cell(row, gi)
            if g:
                carry = g
            via = cell(row, vi)
            if not via or via == "?":
                continue
            rec = {"via": via, "crag": crag, "setor": setor, "tipo": tipo}
            gr = parse_grade(g or carry, prefer=prefer)
            if gr:
                rec["grau"] = gr
            out.append(rec)
    return out


def parse_historicas(ws):
    """Estado / Local / Montanha carry down; running-total rows have no via."""
    out = []
    carry = {0: None, 1: None, 2: None}
    for row in rows_of(ws):
        if is_header_row(row[:4], ["local"]):
            continue
        for i in (0, 1, 2):
            v = cell(row, i)
            if v:
                carry[i] = v
                for j in range(i + 1, 3):   # a new state resets local+montanha
                    carry[j] = None
        via = cell(row, 3)
        if not via:
            continue                        # the accumulated-height rows
        rec = {"via": via, "estado": carry[0], "local": carry[1],
               "montanha": carry[2], "tipo": "via"}
        g = parse_grade(cell(row, 5), prefer="br")
        if g:
            rec["grau"] = g
        y = as_year(cell(row, 4))
        if y:
            rec["ano"] = y
        h = as_height(cell(row, 6))
        if h:
            rec["altura"] = h
        m = cell(row, 7)
        if m:
            rec["material"] = m
        out.append(rec)
    return out


def parse_challenge(ws, cols):
    """big 500 / Croquiteca 1000 / Treino fendas: curated picks of Anhangava vias.
    We keep only (setor, via) so they can flag routes in the main inventory."""
    out = []
    carry = {}
    for row in rows_of(ws):
        if is_header_row(row[:3], ["grau", "gradua"]):
            continue
        gi = cols.get("grau")
        if gi is not None and cell(row, gi):
            carry[gi] = cell(row, gi)
        via = cell(row, cols["via"])
        if not via:
            continue
        rec = {"via": via}
        if "setor" in cols:
            s = cell(row, cols["setor"])
            if s:
                rec["setor"] = s
        if "altura" in cols:
            h = as_height(cell(row, cols["altura"]))
            if h:
                rec["altura"] = h
        out.append(rec)
    return out


def parse_desempenhos(ws):
    """Ascent logs: repeated (Via | Altura | Acumulada) blocks under a climber name."""
    rows = rows_of(ws)
    if len(rows) < 2:
        return []
    names, header = rows[0], rows[1]
    blocks = []
    for i, h in enumerate(header):
        if h and str(h).strip().lower() == "via":
            who = None
            for j in range(i, -1, -1):
                if j < len(names) and names[j]:
                    who = str(names[j]).strip()
                    break
            blocks.append((i, who))
    out = []
    for gi, who in blocks:
        seq = []
        for row in rows[2:]:
            via = cell(row, gi)
            if not via or via.lower() == "via":
                continue
            item = {"via": via}
            h = as_height(cell(row, gi + 1))
            if h:
                item["altura"] = h
            a = as_height(cell(row, gi + 2))
            if a:
                item["acumulada"] = a
            seq.append(item)
        if seq:
            out.append({"quem": who, "vias": seq,
                        "total_m": max((s.get("acumulada") or 0) for s in seq)})
    return out


# --------------------------------------------------------------------------
# geolocation
# --------------------------------------------------------------------------

def band_of(rec):
    """A route's band comes from its free grade, or from the V-scale for boulders."""
    return ((rec.get("grau") or {}).get("band")
            or (rec.get("vscale") or {}).get("band"))


def load_peaks():
    d = json.load(open(PEAKS, encoding="utf-8"))
    by_name = {}
    for f in d["features"]:
        p = f["properties"]
        n = p.get("name")
        if not n:
            continue
        by_name.setdefault(slug(n), []).append(
            {"name": n, "massif": p.get("massif"), "ele": p.get("ele"),
             "coords": f["geometry"]["coordinates"]})
    return by_name


def find_peak(peaks, name, massif=None):
    """Strict when a massif is given. Peak names repeat across the Serra do Mar
    (there is a 'Castelinho' 94 km south of the Anhangava one), so an unconstrained
    first-match fallback silently plants routes on the wrong mountain."""
    cands = peaks.get(slug(name), [])
    if not cands:
        return None
    if massif:
        for c in cands:
            if c.get("massif") and slug(massif) in slug(c["massif"]):
                return c
        return None
    return cands[0] if len(cands) == 1 else None


# Map scope: Serra do Mar / PR mountain crags. Each entry says how to turn a
# route record into the mountain it sits on, and which peak name to look up.
def mountain_of(rec):
    crag = rec.get("crag")
    if crag == "Anhangava":
        return "Anhangava", "Anhangava", "Anhangava-Baitaca"
    if crag == "Castelinhos do Anhangava":
        # The Castelinhos boulder field belongs to the Anhangava complex but has no
        # peak of its own in peaks.geojson (the 'Castelinho' there is a different
        # mountain, in the far south). Rather than invent a coordinate, its problems
        # ride on the Anhangava point and stay separated by `crag` in the popup.
        return "Anhangava", "Anhangava", "Anhangava-Baitaca"
    if crag == "Morro do Canal":
        return "Morro do Canal", "Morro do Canal", "Serra do Canal"
    if crag == "Marumbi":
        # setor reads "Abrolhos - Face Oeste" / "Abrolhos / Torre dos Sinos"
        s = rec.get("setor") or ""
        head = re.split(r"\s*[-/]\s*", s)[0].strip()
        return head or None, head or None, "Marumbi"
    if crag == "Ibitiraquire":
        m = rec.get("montanha")
        return m, m, "Ibitiraquire"
    return None, None, None


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    if not os.path.exists(XLSX):
        raise SystemExit("missing %s" % XLSX)
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    routes = []
    routes += parse_simple(wb["Anhangava"], "Anhangava",
                           {"grau": 0, "via": 1, "setor": 2, "altura": 3, "tecnica": 4})
    routes += parse_simple(wb["Morro do Canal"], "Morro do Canal",
                           {"grau": 0, "via": 1, "setor": 2, "altura": 3})
    routes += parse_simple(wb["Marumbi"], "Marumbi",
                           {"grau": 1, "via": 2, "altura": 3, "setor": 4, "material": 5})
    routes += parse_simple(wb["Ibitiraquire"], "Ibitiraquire",
                           {"grau": 0, "via": 1, "montanha": 2, "altura": 3})
    routes += parse_simple(wb["Boulder Anhangava"], "Anhangava",
                           {"vscale": 0, "grau": 1, "via": 2, "setor": 3},
                           tipo="boulder", ffill=(0, 1))
    # Castelinhos grades on the Hueco V-scale, not the Brazilian one.
    routes += parse_simple(wb["Castelinhos do Anhangava"], "Castelinhos do Anhangava",
                           {"vscale": 0, "via": 1, "bloco": 2, "setor": 3},
                           tipo="boulder", ffill=(0,))
    routes += parse_blocks(wb["São Luis Purunã"], "São Luis Purunã",
                           [(0, 1, "Setor I"), (3, 4, "Setor II"), (6, 7, "Setor III")])
    routes += parse_blocks(wb["Macarrão Ponta Grossa"], "Macarrão Ponta Grossa",
                           [(0, 1, "Frente"), (4, 5, "Fundos")])

    historicas = parse_historicas(wb["Desafio Vias Históricas"])
    desempenhos = parse_desempenhos(wb["Desempenhos Big 1000 e 500"])
    challenges = {
        "big500": parse_challenge(wb["big 500"], {"grau": 0, "via": 1, "setor": 2, "altura": 3}),
        "croquiteca1000": parse_challenge(wb["Croquiteca Raiz 1000"], {"grau": 0, "via": 1, "setor": 2, "altura": 3}),
        "treino_fendas": parse_challenge(wb["Treino fendas Anhangava"], {"grau": 0, "via": 1, "setor": 3, "altura": 4}),
    }

    # The workbook is hand-typed, so the same sector shows up under several
    # spellings ("Abrolhos - Face Oeste" vs "... Face oeste"), which would split
    # one sector into two groups in the popup. Collapse the variants onto the
    # spelling used most often, per crag.
    canon = {}
    for r in routes:
        for key in ("setor", "bloco"):
            if r.get(key):
                canon.setdefault((r["crag"], key, slug(r[key])), Counter())[r[key]] += 1
    merged = 0
    for r in routes:
        for key in ("setor", "bloco"):
            if r.get(key):
                c = canon[(r["crag"], key, slug(r[key]))]
                best = c.most_common(1)[0][0]
                if r[key] != best:
                    r[key] = best
                    merged += 1
    n_variants = sum(1 for c in canon.values() if len(c) > 1)

    # Flag inventory routes that appear in a curated challenge list.
    index = {}
    for r in routes:
        index.setdefault(slug(r["via"]), []).append(r)
    for tag, items in challenges.items():
        hit = 0
        for it in items:
            cands = index.get(slug(it["via"]), [])
            pick = None
            if it.get("setor"):
                for c in cands:
                    if c.get("setor") and slug(c["setor"]) == slug(it["setor"]):
                        pick = c
                        break
            if pick is None and len(cands) == 1:
                pick = cands[0]
            if pick is not None:
                pick.setdefault("listas", []).append(tag)
                hit += 1
        challenges[tag] = {"total": len(items), "casadas": hit, "vias": items}

    # Historic first-ascent years enrich the Marumbi/Ibitiraquire inventory.
    ano_hits = 0
    for h in historicas:
        if h.get("estado") != "Paraná":
            continue
        for c in index.get(slug(h["via"]), []):
            if h.get("ano") and "ano" not in c:
                c["ano"] = h["ano"]
                c.setdefault("listas", []).append("historica")
                ano_hits += 1

    # ---------------- routes.json: everything, unfiltered ----------------
    payload = {
        "fonte": XLSX,
        "escala_bandas": {b: {"label": BAND_LABEL[b], "cor": BAND_COLOR[b]} for b in BANDS},
        "vias": routes,
        "vias_historicas": historicas,
        "desempenhos": desempenhos,
        "listas": challenges,
    }
    json.dump(payload, open(OUT_JSON, "w", encoding="utf-8"),
              ensure_ascii=False, separators=(",", ":"))

    # ---------------- routes.geojson: map scope only ----------------
    peaks = load_peaks()
    groups, unlocated = OrderedDict(), Counter()
    for r in routes:
        mname, peakname, massif = mountain_of(r)
        if not mname:
            continue
        pk = find_peak(peaks, peakname, massif)
        if not pk:
            unlocated[peakname] += 1
            continue
        g = groups.setdefault(mname, {"nome": mname, "massif": pk.get("massif"),
                                      "ele": pk.get("ele"), "coords": pk["coords"],
                                      "vias": []})
        item = {k: v for k, v in r.items() if k != "crag"}
        item["crag"] = r["crag"]
        g["vias"].append(item)

    feats = []
    for mname, g in groups.items():
        vias = g["vias"]
        bands = Counter(b for b in (band_of(v) for v in vias) if b)
        tipos = Counter(v.get("tipo") for v in vias)
        alturas = [v["altura"] for v in vias if v.get("altura")]
        anos = [v["ano"] for v in vias if v.get("ano")]
        hardest, hardest_v = None, None
        for v in vias:
            gr = v.get("grau") or {}
            if gr.get("scale") == "br" and gr.get("free") in BR_SCALE:
                if hardest is None or BR_SCALE.index(gr["free"]) > BR_SCALE.index(hardest):
                    hardest = gr["free"]
            vs = v.get("vscale") or {}
            if vs.get("v") is not None and (hardest_v is None or vs["v"] > hardest_v):
                hardest_v = vs["v"]
        # Dominant band drives the marker colour.
        band = bands.most_common(1)[0][0] if bands else None
        props = {
            "nome": mname,
            "massif": g["massif"],
            "ele": g["ele"],
            "n_vias": tipos.get("via", 0),
            "n_boulders": tipos.get("boulder", 0),
            "total": len(vias),
            "bandas": dict(bands),
            "band": band,
            "color": BAND_COLOR.get(band, "#adb5bd"),
        }
        if hardest:
            props["max_grau"] = hardest
        if hardest_v is not None:
            props["max_boulder"] = "V%d" % hardest_v
        if alturas:
            props["altura_max"] = max(alturas)
        if anos:
            props["ano_min"] = min(anos)
        props["vias"] = sorted(
            vias, key=lambda v: (v.get("tipo") != "via", -(v.get("altura") or 0), v["via"]))
        feats.append({"type": "Feature", "properties": props,
                      "geometry": {"type": "Point", "coordinates": g["coords"]}})

    feats.sort(key=lambda f: -f["properties"]["total"])
    # Band metadata travels with the layer so index.html never redefines the scale.
    bandas_meta = [{"band": b, "label": BAND_LABEL[b], "cor": BAND_COLOR[b],
                    "br": "%s-%s" % BR_BANDS[b], "fr": "%s-%s" % FR_BANDS[b]}
                   for b in BANDS]
    json.dump({"type": "FeatureCollection", "bandas_meta": bandas_meta,
               "fonte": XLSX, "features": feats},
              open(OUT_GEO, "w", encoding="utf-8"), ensure_ascii=False, separators=(",", ":"))

    # ---------------- report ----------------
    print("\n=== VIAS DE ESCALADA ===")
    print("total parseado : %d (%d vias + %d boulders)" % (
        len(routes), sum(1 for r in routes if r.get("tipo") == "via"),
        sum(1 for r in routes if r.get("tipo") == "boulder")))
    print("com grau       : %d (%d escala V)" % (
        sum(1 for r in routes if r.get("grau") or r.get("vscale")),
        sum(1 for r in routes if r.get("vscale"))))
    print("sem faixa      : %d" % sum(1 for r in routes if not band_of(r)))
    print("com altura     : %d" % sum(1 for r in routes if r.get("altura")))
    print("historicas     : %d (%d no PR)" % (
        len(historicas), sum(1 for h in historicas if h.get("estado") == "Paraná")))
    print("ano casado     : %d" % ano_hits)
    print("setores unificados: %d grafias variantes, %d linhas ajustadas" % (n_variants, merged))
    print("desempenhos    : %d sequencias" % len(desempenhos))
    for tag, c in challenges.items():
        print("  lista %-16s %d itens, %d casados" % (tag, c["total"], c["casadas"]))
    print("\npor crag:")
    for c, n in Counter(r["crag"] for r in routes).most_common():
        print("  %-26s %d" % (c, n))
    print("\npor banda:")
    for b, n in Counter(b for b in (band_of(r) for r in routes) if b).most_common():
        print("  %-14s %d" % (b, n))
    print("\nmapa: %d montanhas, %d vias geolocalizadas" % (
        len(feats), sum(f["properties"]["total"] for f in feats)))
    for f in feats:
        p = f["properties"]
        print("  %-20s %3d  max %-8s %-6s %s" % (
            p["nome"], p["total"], p.get("max_grau") or "-",
            p.get("max_boulder") or "-", p["massif"] or ""))
    if unlocated:
        print("\nNAO geolocalizadas:")
        for n, c in unlocated.most_common():
            print("  %-24s %d" % (n, c))
    print("\n%s %.0f KB | %s %.0f KB" % (
        OUT_JSON, os.path.getsize(OUT_JSON) / 1024,
        OUT_GEO, os.path.getsize(OUT_GEO) / 1024))


if __name__ == "__main__":
    main()
